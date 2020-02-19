import pandas as pd
import gdxtools as gt
import datetime
import re
from openpyxl import load_workbook
import glob
import numpy as np


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows[1:], columns=data_rows[0])


def read_elec_demand():
    path = '../../data/raw_data/'
    file = 'table_2-2_load_duration_curves_used_in_epa_platform_v6.xlsx'

    wb = load_workbook(filename=path+file, read_only=True, data_only=True)
    df = load_workbook_range('B4:AB27014', wb['Table 2-2'])

    miso_regions = {'MIS_AMSO', 'MIS_AR', 'MIS_D_MS',
                    'MIS_IA', 'MIS_IL', 'MIS_INKY', 'MIS_LA', 'MIS_LMI', 'MIS_MAPP',
                    'MIS_MIDA', 'MIS_MNWI', 'MIS_MO', 'MIS_WOTA', 'MIS_WUMS'}

    df.drop(df[df['Region'].isin(miso_regions) == False].index, inplace=True)
    df.reset_index(drop=True, inplace=True)

    df = pd.melt(df, id_vars=['Region', 'Month', 'Day'], value_vars=list(
        set(df.keys()) - {'Region', 'Month', 'Day'}), var_name='Hour')

    df.replace({'Hour ': ''}, regex=True, inplace=True)
    df['Month'] = df['Month'].map(int)
    df['Day'] = df['Day'].map(int)
    df['Hour'] = df['Hour'].map(int)

    df['timestamp'] = [datetime.datetime(year=2021, month=1, day=1) + datetime.timedelta(
        days=int(df.loc[n, 'Day'])-1, hours=int(df.loc[n, 'Hour'])-1) for n, v in enumerate(df.value)]

    df['epoch'] = [df.loc[n, 'timestamp'].timestamp() for n, v in enumerate(df.value)]

    df['hour'] = [(df.loc[n, 'timestamp'] - datetime.datetime(year=2021, month=1,
                                                              day=1)).total_seconds() / 60 / 60 + 1 for n, v in enumerate(df.value)]

    df['epoch'] = df['epoch'].map(int)
    df['hour'] = df['hour'].map(int)

    df = df[['timestamp', 'epoch', 'hour', 'Region', 'value']].copy()
    df.sort_values(by=['Region', 'epoch'], ascending=True, inplace=True)
    df.reset_index(drop=True, inplace=True)

    # map in day of week
    # monday is 0 and sunday is 6
    df['weekday'] = [i.weekday() for i in df['timestamp']]
    day = {0: 'weekday', 1: 'weekday', 2: 'weekday',
           3: 'weekday', 4: 'weekday', 5: 'weekend', 6: 'weekend'}
    df['daytype'] = df['weekday'].map(day)

    # 24 hour label
    df['24hr'] = [i.hour for i in df['timestamp']]

    return df


def read_lmp():
    path = '../../data/raw_data/miso_historical_lmp/'

    files = glob.glob(path+'*.csv')

    df = pd.DataFrame(columns=['MARKET_DAY', 'NODE', 'TYPE', 'VALUE', 'HE01', 'HE02', 'HE03', 'HE04', 'HE05', 'HE06', 'HE07', 'HE08', 'HE09',
                               'HE10', 'HE11', 'HE12', 'HE13', 'HE14', 'HE15', 'HE16', 'HE17', 'HE18', 'HE19', 'HE20', 'HE21', 'HE22', 'HE23', 'HE24'])

    relabel = {'HE1': 'HE01', 'HE2': 'HE02', 'HE3': 'HE03', 'HE4': 'HE04',
               'HE5': 'HE05', 'HE6': 'HE06', 'HE7': 'HE07', 'HE8': 'HE08', 'HE9': 'HE09'}

    hrs = ['HE01', 'HE02', 'HE03', 'HE04', 'HE05', 'HE06', 'HE07', 'HE08', 'HE09',
           'HE10', 'HE11', 'HE12', 'HE13', 'HE14', 'HE15', 'HE16', 'HE17', 'HE18', 'HE19', 'HE20', 'HE21', 'HE22', 'HE23', 'HE24']

    wi_lmp = ['WEC.OKCGC8', 'WEC.OKCGC7', 'WEC.OAKCREKC5', 'WEC.ERG2', 'WEC.CC.PORTW2', 'WEC.CC.PORTW1', 'ALTE.EDGG5G5', 'WEC.S', 'WEC.PTBHGB2', 'WEC.PTBHGB1', 'WPS.COLUMBIA2',
              'WPS.COLUMBIA1', 'ALTE.COLUMBAL2', 'ALTE.COLUMBAL1', 'ALTW.LANSIN4', 'DPC.GENOA3', 'DPC.DPC', 'DPC.JPM', 'WPS.WESTON4', 'WPS.WESTON3', 'WPS.DPC.WESTN4', 'WPS.WPSM', 'NSP.KING1']

    hub_lmp = ['MINN.HUB', 'ILLINOIS.HUB', 'MICHIGAN.HUB']

    all_lmp = wi_lmp.copy()
    all_lmp.extend(hub_lmp)

    n = 0
    for i in files:
        n += 1
        print('Processing... ' + i + ', ' + str(n) + ' of ' + str(len(files)) + ' files')
        t = pd.read_csv(i, index_col=None, engine='c')
        t.rename(columns=relabel, inplace=True)

        t.drop(index=t[t['VALUE'] != 'LMP'].index, inplace=True)
        t.drop(index=t[t['NODE'].isin(all_lmp) == False].index, inplace=True)

        df = df.append(t, sort=True, ignore_index=True)

    # convert all hour data to float
    for i in hrs:
        df[i] = df[i].map(float)

    df = df[['MARKET_DAY', 'NODE', 'TYPE', 'VALUE', 'HE01', 'HE02', 'HE03', 'HE04', 'HE05', 'HE06', 'HE07', 'HE08', 'HE09',
             'HE10', 'HE11', 'HE12', 'HE13', 'HE14', 'HE15', 'HE16', 'HE17', 'HE18', 'HE19', 'HE20', 'HE21', 'HE22', 'HE23', 'HE24']].copy()

    df['MARKET_DAY'] = pd.to_datetime(df['MARKET_DAY'], format='%m/%d/%Y')
    df.drop(columns='VALUE', inplace=True)

    df = pd.melt(df, id_vars=['MARKET_DAY', 'NODE', 'TYPE'], var_name='hour', value_name='LMP')
    df['hour'] = [df.loc[i, 'hour'].split('HE')[1] for i in df.index]
    df['hour'] = df.hour.map(int)

    df['MARKET_DAY'] = [df.loc[i, 'MARKET_DAY'] +
                        datetime.timedelta(hours=df.loc[i, 'hour'].item()) - datetime.timedelta(seconds=1) for i in df.index]

    df.drop(columns=['hour', 'TYPE'], inplace=True)

    df2 = pd.pivot_table(df[df['NODE'].isin(wi_lmp) == True], values=[
                         'LMP'], index=['MARKET_DAY'], aggfunc=np.mean)

    df2.rename(columns={'LMP': 'WISCONSIN.HUB'}, inplace=True)

    for i in hub_lmp:
        df2[i] = df[df['NODE'] == i].set_index('MARKET_DAY').LMP

    df2.reset_index(drop=False, inplace=True)

    return df2


if __name__ == '__main__':

    elec = read_elec_demand()
    lmp = read_lmp()

    gdxout = gt.gdxrw.gdx_writer('lmp_data.gdx')

    gdxout.add_set(gamssetname='hr', toset=list(range(1, 8761)), desc='hour')
    gdxout.add_set(gamssetname='i', toset=['WI', 'MN', 'IL', 'MI'], desc='regions')

    lmp_2018 = lmp[lmp['MARKET_DAY'].dt.year == 2018].copy()
    lmp_2018.reset_index(drop=True, inplace=True)
    lmp_2018.reset_index(drop=False, inplace=True)
    lmp_2018['index'] = lmp_2018['index'] + 1
    lmp_2018.drop(columns=['MARKET_DAY'], inplace=True)

    p = pd.melt(lmp_2018, id_vars='index')
    p['variable'] = p['variable'].map(
        {'WISCONSIN.HUB': 'WI', 'MINN.HUB': 'MN', 'ILLINOIS.HUB': 'IL', 'MICHIGAN.HUB': 'MI'})

    p = dict(zip(zip(p['index'], p['variable']), p.value))

    gdxout.add_parameter_dc(gamsparametername='lmp', domain=[
        'hr', 'i'], toparameter=p, desc='LMP price (units: $/MW)')

    # fake map between LMP regions and Load curve data
    m = {'MIS_WUMS': 'WI', 'MIS_MNWI': 'MN', 'MIS_IL': 'IL', 'MIS_MIDA': 'MI'}
    elec['Region'] = elec['Region'].map(m)
    elec['test'] = elec['Region'].isin(list(m.values()))

    elec.drop(index=elec[elec['test'] == False].index, inplace=True)

    e = dict(zip(zip(elec.hour, elec.Region), elec.value))
    gdxout.add_parameter_dc(gamsparametername='load', domain=[
                            'hr', 'i'], toparameter=e, desc='Load (units: MW)')

    gdxout.export_gdx()
