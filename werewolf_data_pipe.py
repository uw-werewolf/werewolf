import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re
import datetime
import folium
import pickle
import gmsxfr
import filesys as fs
import os


def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows[1:], columns=data_rows[0])


def read_fips_region(to_csv=False):
    filename = "v 6.17 IPM Region-State-County Mapping - 04-03-17 v2 ToEPA.xlsx"

    wb = load_workbook(
        filename=os.path.join(fs.raw_data_dir, filename), read_only=True, data_only=True
    )

    df = load_workbook_range("A1:D3110", wb["TO_GIS 04-03-2017"])
    df.rename(
        columns={
            "FIPS5": "fips",
            "State": "state",
            "County": "county",
            "v6.17 Model Region": "Region Name",
        },
        inplace=True,
    )

    mapping = pd.read_csv(os.path.join(fs.map_dir, "region_codes.csv"), index_col=None)
    df["state"] = df["state"].map(dict(zip(mapping["from"], mapping["to"])))

    df["fips"] = df["fips"].map(int).map(str)

    if to_csv == True:
        df.to_csv(os.path.join(fs.processed_data_dir, "fips_region_mapping.csv"))

    return df


def read_needs_data(to_csv=False):
    filename = "needs_v6_09-30-19.xlsx"

    wb = load_workbook(
        filename=os.path.join(fs.raw_data_dir, filename), read_only=True, data_only=True
    )
    df = load_workbook_range("A1:AX18564", wb["NEEDS v6_active"])

    # pull out only these columns
    cols = [
        "Plant Name",
        "UniqueID_Final",
        "ORIS Plant Code",
        "Boiler/Generator/Committed Unit",
        "Unit ID",
        "CAMD Database UnitID",
        "PlantType",
        "Combustion Turbine/IC Engine",
        "Region Name",
        "State Code",
        "County",
        "County Code",
        "FIPS5",
        "Capacity (MW)",
        "Heat Rate (Btu/kWh)",
        "On Line Year",
        "Retirement Year",
        "Firing",
        "Bottom",
        "Cogen?",
        "Modeled Fuels",
    ]

    df = df[cols].copy()

    # clean up
    df.replace(
        {"#": "", " ": "_", "/": "", "\(": "", "\)": "", "-": "_"},
        regex=True,
        inplace=True,
    )

    df["FIPS5"] = df["FIPS5"].map(int).map(str)
    df["State Code"] = df["State Code"].map(str)
    df["County Code"] = df["County Code"].map(int)
    df["ORIS Plant Code"] = df["ORIS Plant Code"].map(int)
    df["On Line Year"] = df["On Line Year"].map(int)
    df["Retirement Year"] = df["Retirement Year"].map(int)

    mapping = pd.read_csv(os.path.join(fs.map_dir, "region_codes.csv"), index_col=None)
    df["State Code"] = df["State Code"].map(dict(zip(mapping["from"], mapping["to"])))

    if to_csv == True:
        df.to_csv(os.path.join(fs.processed_data_dir, "epa_needs_data.csv"))

    return df


# def read_egrid_data(to_csv=False):
#     path = './raw_data/'
#     file = 'egrid2016_data.xlsx'
#
#     wb = load_workbook(filename=path+file, read_only=True, data_only=True)
#     df = load_workbook_range('A2:P26185', wb['GEN16'])
#
#     if to_csv == True:
#         df.to_csv('./processed_data/epa_egrid_data.csv')
#
#     return df


# def read_solar(to_csv=False):
#     path = './raw_data/'
#     file = 'solarsummaries.xlsx'
#
#     wb = load_workbook(filename=path+file, read_only=True, data_only=True)
#     df = load_workbook_range('A1:BP3142', wb['GHI County'])
#
#     df = pd.melt(df, id_vars=['County', 'County FIPS'], value_vars=list(
#         set(df.keys()) - {'State', 'County', 'County FIPS'}))
#
#     if to_csv == True:
#         df.to_csv('./processed_data/solar_cnty_ghi_data.csv')
#
#     return df


def read_latlng(to_csv=False):
    filename = "population_geo.csv"

    df = pd.read_csv(os.path.join(fs.raw_data_dir, filename), index_col=None)
    df["STATEFP"] = df["STATEFP"] * 1000
    df["fips"] = df["STATEFP"] + df["COUNTYFP"]
    df["fips"] = df["fips"].map(str)

    mapping = pd.read_csv(os.path.join(fs.map_dir, "region_codes.csv"), index_col=None)
    df["STNAME"] = df["STNAME"].map(dict(zip(mapping["from"], mapping["to"])))

    df = df[["fips", "STNAME", "POPULATION", "LATITUDE", "LONGITUDE"]].copy()
    df.rename(
        columns={
            "STNAME": "state",
            "POPULATION": "population",
            "LATITUDE": "lat",
            "LONGITUDE": "lng",
        },
        inplace=True,
    )

    if to_csv == True:
        df.to_csv(os.path.join(fs.processed_data_dir, "geo_latlng.csv"))

    return df


def read_elec_demand(to_csv=False):
    filename = "table_2-2_load_duration_curves_used_in_epa_platform_v6.xlsx"

    wb = load_workbook(
        filename=os.path.join(fs.raw_data_dir, filename), read_only=True, data_only=True
    )
    df = load_workbook_range("B4:AB27014", wb["Table 2-2"])

    df = pd.melt(
        df,
        id_vars=["Region", "Month", "Day"],
        value_vars=list(set(df.keys()) - {"Region", "Month", "Day"}),
        var_name="Hour",
    )

    df.replace({"Hour ": ""}, regex=True, inplace=True)
    df["Month"] = df["Month"].map(int)
    df["Day"] = df["Day"].map(int)
    df["Hour"] = df["Hour"].map(int)

    df["timestamp"] = [
        datetime.datetime(year=2021, month=1, day=1)
        + datetime.timedelta(
            days=int(df.loc[n, "Day"]) - 1, hours=int(df.loc[n, "Hour"]) - 1
        )
        for n, v in enumerate(df.value)
    ]

    df["epoch"] = [df.loc[n, "timestamp"].timestamp() for n, v in enumerate(df.value)]

    df["hour"] = [
        (
            df.loc[n, "timestamp"] - datetime.datetime(year=2021, month=1, day=1)
        ).total_seconds()
        / 60
        / 60
        + 1
        for n, v in enumerate(df.value)
    ]

    df["epoch"] = df["epoch"].map(int)
    df["hour"] = df["hour"].map(int)

    df = df[["timestamp", "epoch", "hour", "Region", "value"]].copy()
    df.sort_values(by=["Region", "epoch"], ascending=True, inplace=True)
    df.reset_index(drop=True, inplace=True)

    df.replace(
        {"NY_Z_G-I": "NY_Z_G_I"}, inplace=True
    )  # fix typo to harmonize with other data

    if to_csv == True:
        df.to_csv(os.path.join(fs.processed_data_dir, "elec_demand.csv"))
        df[["timestamp", "epoch", "hour"]].to_csv(
            os.path.join(fs.processed_data_dir, "time_mapping.csv")
        )

    return df


def read_nrel_reeds_wind_data(to_csv=False):
    file = [
        "LDC_static_inputs_06_27_2019_wind1.pkl",
        "LDC_static_inputs_06_27_2019_wind2.pkl",
        "LDC_static_inputs_06_27_2019_wind3.pkl",
        "LDC_static_inputs_06_27_2019_wind4.pkl",
    ]

    data = []
    for i in file:
        with open(os.path.join(fs.raw_data_dir, i), "rb") as f:
            d = pickle.load(f)
        data.append(d)

    df = pd.concat(data, axis=1)
    df.reset_index(drop=False, inplace=True)

    df = pd.melt(df, id_vars=["hour"])
    new = df["resource"].str.split("_", expand=True)

    df["tech"] = new[0]
    df["trb"] = new[1]
    df["reeds_region"] = new[2]
    df.drop(columns="resource", inplace=True)

    df = df[["tech", "trb", "reeds_region", "hour", "value"]]
    df["tech"] = df["tech"].map(
        {"wind-ons": "Onshore_Wind", "wind-ofs": "Offshore_Wind"}
    )
    df["tech"] = df["tech"] + "_" + df["trb"]

    # add in timestamp congruent with load duration curve time
    time = pd.read_csv(
        os.path.join(fs.processed_data_dir, "time_mapping.csv"), index_col=0
    )

    df["timestamp"] = df["hour"].map(dict(zip(time["hour"], time["timestamp"])))

    if to_csv == True:
        df.to_csv(
            os.path.join(fs.processed_data_dir, "wind_hourly_capacity_factor.csv")
        )

    return df


def read_nrel_reeds_solar_data(to_csv=False):
    file = ["LDC_static_inputs_06_27_2019_upv.pkl"]

    data = []
    for i in file:
        with open(os.path.join(fs.raw_data_dir, i), "rb") as f:
            d = pickle.load(f)
        data.append(d)

    df = pd.concat(data, axis=1)
    df.reset_index(drop=False, inplace=True)

    df = pd.melt(df, id_vars=["hour"])
    new = df["resource"].str.split("_", expand=True)

    df["tech"] = new[0]
    df["trb"] = new[1]
    df["reeds_region"] = new[2]
    df.drop(columns="resource", inplace=True)

    df = df[["tech", "trb", "reeds_region", "hour", "value"]]
    df["tech"] = df["tech"].map({"dupv": "SolarDistUtil_PV", "upv": "SolarUtil_PV"})
    df["tech"] = df["tech"] + "_" + df["trb"]

    # add in timestamp congruent with load duration curve time
    time = pd.read_csv(
        os.path.join(fs.processed_data_dir, "time_mapping.csv"), index_col=0
    )

    df["timestamp"] = df["hour"].map(dict(zip(time["hour"], time["timestamp"])))

    if to_csv == True:
        df.to_csv(
            os.path.join(fs.processed_data_dir, "solar_hourly_capacity_factor.csv")
        )

    return df


def read_reeds_regions(to_csv=False):
    filename = "nrel_reeds_region_mapping.csv"

    reeds_reg = pd.read_csv(os.path.join(fs.raw_data_dir, filename), index_col=0)
    reeds_reg["fips"] = reeds_reg["state.fips"] * 1000 + reeds_reg["cnty.fips"]
    reeds_reg.drop(
        reeds_reg[reeds_reg["state.abb"].isin({"HI", "AK"}) == True].index, inplace=True
    )

    # fix NREL ReEDS regions (errors?)
    reeds_reg.loc[reeds_reg[reeds_reg["fips"] == 37133].index, "reeds.reg"] = 293
    reeds_reg.loc[reeds_reg[reeds_reg["fips"] == 53029].index, "reeds.reg"] = 3

    reeds_reg["reeds.reg"] = reeds_reg["reeds.reg"].map(int)

    reeds_reg["reeds.reg"] = ["s" + str(i) for i in reeds_reg["reeds.reg"]]
    reeds_reg["reeds.ba"] = ["p" + str(i) for i in reeds_reg["reeds.ba"]]

    if to_csv == True:
        reeds_reg.to_csv(os.path.join(fs.processed_data_dir, "nrel_regions.csv"))

    return reeds_reg


def read_reeds_wind_cap(to_csv=False):
    filename = "wind_supply_curves_capacity_NARIS.csv"

    reeds_reg = pd.read_csv(
        os.path.join(fs.processed_data_dir, "nrel_regions.csv"), index_col=None
    )

    df = pd.read_csv(os.path.join(fs.raw_data_dir, filename), index_col=None)
    df.rename(
        columns={"Unnamed: 0": "reeds.reg", "Unnamed: 1": "trb", "Unnamed: 2": "tech"},
        inplace=True,
    )

    df.rename(
        columns={
            "wsc1": "cost_bin1",
            "wsc2": "cost_bin2",
            "wsc3": "cost_bin3",
            "wsc4": "cost_bin4",
            "wsc5": "cost_bin5",
        },
        inplace=True,
    )

    # add in leading 's' to regions for consistancy
    df["reeds.reg"] = df["reeds.reg"].map(str)
    df["reeds.reg"] = "s" + df["reeds.reg"]

    # drop regions that are not in nrel_regions.csv file
    df.drop(
        df[df["reeds.reg"].isin(reeds_reg["reeds.reg"]) == False].index, inplace=True
    )
    df.reset_index(drop=True, inplace=True)

    df["trb"] = [i[1] for i in df["trb"].str.split("class")]

    df["tech"] = df["tech"].map(
        {"wind-ons": "Onshore_Wind", "wind-ofs": "Offshore_Wind"}
    )
    df["tech"] = df["tech"] + "_" + df["trb"]

    if to_csv == True:
        df.to_csv(os.path.join(fs.processed_data_dir, "wind_capacity_MW.csv"))

    return df


def read_reeds_wind_costs(to_csv=False):
    filename = "wind_supply_curves_cost_NARIS.csv"

    reeds_reg = pd.read_csv(
        os.path.join(fs.processed_data_dir, "nrel_regions.csv"), index_col=None
    )

    df = pd.read_csv(os.path.join(fs.raw_data_dir, filename), index_col=None)
    df.rename(
        columns={"Unnamed: 0": "reeds.reg", "Unnamed: 1": "trb", "Unnamed: 2": "tech"},
        inplace=True,
    )

    # add in leading 's' to regions for consistancy
    df["reeds.reg"] = df["reeds.reg"].map(str)
    df["reeds.reg"] = "s" + df["reeds.reg"]

    # drop regions that are not in nrel_regions.csv file
    df.drop(
        df[df["reeds.reg"].isin(reeds_reg["reeds.reg"]) == False].index, inplace=True
    )
    df.reset_index(drop=True, inplace=True)

    df["trb"] = [i[1] for i in df["trb"].str.split("class")]

    df["tech"] = df["tech"].map(
        {"wind-ons": "Onshore_Wind", "wind-ofs": "Offshore_Wind"}
    )
    df["tech"] = df["tech"] + "_" + df["trb"]

    df.rename(
        columns={
            "wsc1": "cost_bin1",
            "wsc2": "cost_bin2",
            "wsc3": "cost_bin3",
            "wsc4": "cost_bin4",
            "wsc5": "cost_bin5",
        },
        inplace=True,
    )

    if to_csv == True:
        df.to_csv(os.path.join(fs.processed_data_dir, "wind_dollar_per_MW.csv"))

    return df


def read_reeds_solar_cap(to_csv=False):
    file = "UPV_supply_curves_capacity_NARIS.csv"
    file2 = "DUPV_supply_curves_capacity_NARIS.csv"

    reeds_reg = pd.read_csv(
        os.path.join(fs.processed_data_dir, "nrel_regions.csv"), index_col=None
    )

    df = pd.read_csv(os.path.join(fs.raw_data_dir, file), index_col=None)
    df["tech"] = "SolarUtil_PV"

    df2 = pd.read_csv(os.path.join(fs.raw_data_dir, file2), index_col=None)
    df2["tech"] = "SolarDistUtil_PV"

    df.rename(
        columns={
            "upvsc1": "cost_bin1",
            "upvsc2": "cost_bin2",
            "upvsc3": "cost_bin3",
            "upvsc4": "cost_bin4",
            "upvsc5": "cost_bin5",
        },
        inplace=True,
    )

    df2.rename(
        columns={
            "dupvsc1": "cost_bin1",
            "dupvsc2": "cost_bin2",
            "dupvsc3": "cost_bin3",
            "dupvsc4": "cost_bin4",
            "dupvsc5": "cost_bin5",
        },
        inplace=True,
    )

    df.rename(columns={"Unnamed: 0": "reeds.ba", "Unnamed: 1": "trb"}, inplace=True)
    df2.rename(columns={"Unnamed: 0": "reeds.ba", "Unnamed: 1": "trb"}, inplace=True)

    # reorder the columns
    df = df[
        [
            "reeds.ba",
            "trb",
            "tech",
            "cost_bin1",
            "cost_bin2",
            "cost_bin3",
            "cost_bin4",
            "cost_bin5",
        ]
    ].copy()
    df2 = df2[
        [
            "reeds.ba",
            "trb",
            "tech",
            "cost_bin1",
            "cost_bin2",
            "cost_bin3",
            "cost_bin4",
            "cost_bin5",
        ]
    ].copy()

    df = pd.concat([df, df2], axis=0)

    df["trb"] = [i[1] for i in df["trb"].str.split("class")]
    df["tech"] = df["tech"] + "_" + df["trb"]

    # drop regions that are not in nrel_regions.csv file
    df.drop(df[df["reeds.ba"].isin(reeds_reg["reeds.ba"]) == False].index, inplace=True)
    df.reset_index(drop=True, inplace=True)

    if to_csv == True:
        df.to_csv(os.path.join(fs.processed_data_dir, "solar_capacity_MW.csv"))

    return df


def read_reeds_solar_costs(to_csv=False):
    file = "UPV_supply_curves_cost_NARIS.csv"
    file2 = "DUPV_supply_curves_cost_NARIS.csv"

    reeds_reg = pd.read_csv(
        os.path.join(fs.processed_data_dir, "nrel_regions.csv"), index_col=None
    )

    df = pd.read_csv(os.path.join(fs.raw_data_dir, file), index_col=None)
    df["tech"] = "SolarUtil_PV"

    df2 = pd.read_csv(os.path.join(fs.raw_data_dir, file2), index_col=None)
    df2["tech"] = "SolarDistUtil_PV"

    df.rename(
        columns={
            "upvsc1": "cost_bin1",
            "upvsc2": "cost_bin2",
            "upvsc3": "cost_bin3",
            "upvsc4": "cost_bin4",
            "upvsc5": "cost_bin5",
        },
        inplace=True,
    )

    df2.rename(
        columns={
            "dupvsc1": "cost_bin1",
            "dupvsc2": "cost_bin2",
            "dupvsc3": "cost_bin3",
            "dupvsc4": "cost_bin4",
            "dupvsc5": "cost_bin5",
        },
        inplace=True,
    )

    df.rename(columns={"Unnamed: 0": "reeds.ba", "Unnamed: 1": "trb"}, inplace=True)
    df2.rename(columns={"Unnamed: 0": "reeds.ba", "Unnamed: 1": "trb"}, inplace=True)

    # reorder the columns
    df = df[
        [
            "reeds.ba",
            "trb",
            "tech",
            "cost_bin1",
            "cost_bin2",
            "cost_bin3",
            "cost_bin4",
            "cost_bin5",
        ]
    ].copy()
    df2 = df2[
        [
            "reeds.ba",
            "trb",
            "tech",
            "cost_bin1",
            "cost_bin2",
            "cost_bin3",
            "cost_bin4",
            "cost_bin5",
        ]
    ].copy()

    df = pd.concat([df, df2], axis=0)

    df["trb"] = [i[1] for i in df["trb"].str.split("class")]
    df["tech"] = df["tech"] + "_" + df["trb"]

    # drop regions that are not in nrel_regions.csv file
    df.drop(df[df["reeds.ba"].isin(reeds_reg["reeds.ba"]) == False].index, inplace=True)
    df.reset_index(drop=True, inplace=True)

    if to_csv == True:
        df.to_csv(os.path.join(fs.processed_data_dir, "solar_dollar_per_MW.csv"))

    return df


if __name__ == "__main__":

    # load duration curves
    load = read_elec_demand(to_csv=True)
    cn_regions = [
        "CN_AB",
        "CN_BC",
        "CN_MB",
        "CN_NB",
        "CN_NF",
        "CN_NL",
        "CN_NS",
        "CN_ON",
        "CN_PE",
        "CN_PQ",
        "CN_SK",
    ]
    load.drop(
        load[load.Region.isin(cn_regions) == True].index, inplace=True
    )  # drop Canadian regions
    load.reset_index(drop=True, inplace=True)

    # read in generator data
    epa_needs = read_needs_data(to_csv=True)
    epa_needs.replace({"Energy_Storage": "Battery_slow"}, inplace=True)

    # read in fips:region mapping data
    fips_region = read_fips_region(to_csv=True)
    fips_region.replace({"NY_Z_G-I": "NY_Z_G_I"}, inplace=True)  # cleanup of a typo

    # pull in all regions
    latlng = read_latlng(to_csv=True)
    latlng.drop(
        latlng[latlng["state"].isin(["PR", "AK", "HI"]) == True].index, inplace=True
    )  # drop PR, AK, and HI from data
    latlng.reset_index(drop=True, inplace=True)  # reset index
    latlng["fips"] = (
        latlng["fips"].map(int).map(str)
    )  # gets rid of leading zeros in fips

    latlng["Region Name"] = latlng["fips"].map(
        dict(zip(fips_region["fips"], fips_region["Region Name"]))
    )

    reeds_reg = read_reeds_regions(to_csv=True)

    nrel_wind_cap = read_reeds_wind_cap(to_csv=True)
    nrel_solar_cap = read_reeds_solar_cap(to_csv=True)

    nrel_wind_cost = read_reeds_wind_costs(to_csv=True)
    nrel_solar_cost = read_reeds_solar_costs(to_csv=True)

    # egrid = read_egrid_data(to_csv=True)
    # solar = read_solar(to_csv=True)

    # m = epa_needs[n].merge(egrid[['ORISPL', 'GENID', 'NAMEPCAP', 'CFACT']], how='outer', left_on=[
    #     'ORIS Plant Code', 'Unit ID'], right_on=['ORISPL', 'GENID'], indicator=True)
    #
    # cols = ['Plant Name', 'UniqueID_Final', 'ORIS Plant Code', 'Boiler/Generator/Committed Unit', 'Unit ID', 'CAMD Database UnitID', 'PlantType', 'Combustion Turbine/IC Engine', 'Region Name',
    #         'State Name', 'State Code', 'County', 'County Code', 'FIPS5', 'On Line Year', 'Retirement Year', 'Firing', 'Bottom', 'Cogen?', 'Modeled Fuels', 'Capacity (MW)', 'Heat Rate (Btu/kWh)', 'CFACT']
    #
    # m = m[cols].copy()
    # m.rename(columns={'CFACT': 'Capacity Factor'}, inplace=True)
    # m = m[m['State Name'] == 'Wisconsin'].copy()
    #
    # m.to_csv('./processed_data/wisconsin_generators.csv')

    k = epa_needs["PlantType"].unique().tolist()
    k.extend(["Battery_med", "Battery_fast"])
    k.extend(list(set(nrel_wind_cap.tech)))
    k.extend(list(set(nrel_solar_cap.tech)))
    k.sort()

    gen = [
        "Biomass",
        "Coal_Steam",
        "Combined_Cycle",
        "Combustion_Turbine",
        "Fossil_Waste",
        "Fuel_Cell",
        "Geothermal",
        "Hydro",
        "IGCC",
        "Landfill_Gas",
        "Municipal_Solid_Waste",
        "Non_Fossil_Waste",
        "Nuclear",
        "OG_Steam",
        "Offshore_Wind",
        "Offshore_Wind_1",
        "Offshore_Wind_10",
        "Offshore_Wind_11",
        "Offshore_Wind_12",
        "Offshore_Wind_13",
        "Offshore_Wind_14",
        "Offshore_Wind_15",
        "Offshore_Wind_2",
        "Offshore_Wind_3",
        "Offshore_Wind_4",
        "Offshore_Wind_5",
        "Offshore_Wind_6",
        "Offshore_Wind_7",
        "Offshore_Wind_8",
        "Offshore_Wind_9",
        "Onshore_Wind",
        "Onshore_Wind_1",
        "Onshore_Wind_10",
        "Onshore_Wind_2",
        "Onshore_Wind_3",
        "Onshore_Wind_4",
        "Onshore_Wind_5",
        "Onshore_Wind_6",
        "Onshore_Wind_7",
        "Onshore_Wind_8",
        "Onshore_Wind_9",
        "Pumped_Storage",
        "SolarDistUtil_PV_1",
        "SolarDistUtil_PV_2",
        "SolarDistUtil_PV_3",
        "SolarDistUtil_PV_4",
        "SolarDistUtil_PV_5",
        "SolarDistUtil_PV_6",
        "SolarDistUtil_PV_7",
        "SolarUtil_PV_1",
        "SolarUtil_PV_2",
        "SolarUtil_PV_3",
        "SolarUtil_PV_4",
        "SolarUtil_PV_5",
        "SolarUtil_PV_6",
        "SolarUtil_PV_7",
        "Solar_PV",
        "Solar_Thermal",
        "Tires",
    ]

    fossil = [
        "Coal_Steam",
        "Combined_Cycle",
        "Combustion_Turbine",
        "Fossil_Waste",
        "IGCC",
        "Landfill_Gas",
        "Municipal_Solid_Waste",
        "Non_Fossil_Waste",
        "OG_Steam",
        "Tires",
    ]

    hydro = ["Hydro", "Pumped_Storage"]
    geothermal = ["Geothermal"]
    nuclear = ["Nuclear"]
    store = ["Battery_fast", "Battery_med", "Battery_slow", "Pumped_Storage"]
    battery = ["Battery_fast", "Battery_med", "Battery_slow"]
    renew = [
        "Battery_fast",
        "Battery_med",
        "Battery_slow",
        "Fuel_Cell",
        "Hydro",
        "Offshore_Wind",
        "Offshore_Wind_1",
        "Offshore_Wind_10",
        "Offshore_Wind_11",
        "Offshore_Wind_12",
        "Offshore_Wind_13",
        "Offshore_Wind_14",
        "Offshore_Wind_15",
        "Offshore_Wind_2",
        "Offshore_Wind_3",
        "Offshore_Wind_4",
        "Offshore_Wind_5",
        "Offshore_Wind_6",
        "Offshore_Wind_7",
        "Offshore_Wind_8",
        "Offshore_Wind_9",
        "Onshore_Wind",
        "Onshore_Wind_1",
        "Onshore_Wind_10",
        "Onshore_Wind_2",
        "Onshore_Wind_3",
        "Onshore_Wind_4",
        "Onshore_Wind_5",
        "Onshore_Wind_6",
        "Onshore_Wind_7",
        "Onshore_Wind_8",
        "Onshore_Wind_9",
        "Pumped_Storage",
        "SolarDistUtil_PV_1",
        "SolarDistUtil_PV_2",
        "SolarDistUtil_PV_3",
        "SolarDistUtil_PV_4",
        "SolarDistUtil_PV_5",
        "SolarDistUtil_PV_6",
        "SolarDistUtil_PV_7",
        "SolarUtil_PV_1",
        "SolarUtil_PV_2",
        "SolarUtil_PV_3",
        "SolarUtil_PV_4",
        "SolarUtil_PV_5",
        "SolarUtil_PV_6",
        "SolarUtil_PV_7",
        "Solar_PV",
        "Solar_Thermal",
    ]

    all_solar = [
        "SolarDistUtil_PV_1",
        "SolarDistUtil_PV_2",
        "SolarDistUtil_PV_3",
        "SolarDistUtil_PV_4",
        "SolarDistUtil_PV_5",
        "SolarDistUtil_PV_6",
        "SolarDistUtil_PV_7",
        "SolarUtil_PV_1",
        "SolarUtil_PV_2",
        "SolarUtil_PV_3",
        "SolarUtil_PV_4",
        "SolarUtil_PV_5",
        "SolarUtil_PV_6",
        "SolarUtil_PV_7",
        "Solar_PV",
        "Solar_Thermal",
    ]

    all_solar_PV = [
        "SolarDistUtil_PV_1",
        "SolarDistUtil_PV_2",
        "SolarDistUtil_PV_3",
        "SolarDistUtil_PV_4",
        "SolarDistUtil_PV_5",
        "SolarDistUtil_PV_6",
        "SolarDistUtil_PV_7",
        "SolarUtil_PV_1",
        "SolarUtil_PV_2",
        "SolarUtil_PV_3",
        "SolarUtil_PV_4",
        "SolarUtil_PV_5",
        "SolarUtil_PV_6",
        "SolarUtil_PV_7",
        "Solar_PV",
    ]

    all_distsolar_PV = [
        "SolarDistUtil_PV_1",
        "SolarDistUtil_PV_2",
        "SolarDistUtil_PV_3",
        "SolarDistUtil_PV_4",
        "SolarDistUtil_PV_5",
        "SolarDistUtil_PV_6",
        "SolarDistUtil_PV_7",
    ]

    all_utilsolar_PV = [
        "SolarUtil_PV_1",
        "SolarUtil_PV_2",
        "SolarUtil_PV_3",
        "SolarUtil_PV_4",
        "SolarUtil_PV_5",
        "SolarUtil_PV_6",
        "SolarUtil_PV_7",
        "Solar_PV",
    ]

    all_solar_therm = ["Solar_Thermal"]

    all_wind = [
        "Offshore_Wind",
        "Offshore_Wind_1",
        "Offshore_Wind_10",
        "Offshore_Wind_11",
        "Offshore_Wind_12",
        "Offshore_Wind_13",
        "Offshore_Wind_14",
        "Offshore_Wind_15",
        "Offshore_Wind_2",
        "Offshore_Wind_3",
        "Offshore_Wind_4",
        "Offshore_Wind_5",
        "Offshore_Wind_6",
        "Offshore_Wind_7",
        "Offshore_Wind_8",
        "Offshore_Wind_9",
        "Onshore_Wind",
        "Onshore_Wind_1",
        "Onshore_Wind_10",
        "Onshore_Wind_2",
        "Onshore_Wind_3",
        "Onshore_Wind_4",
        "Onshore_Wind_5",
        "Onshore_Wind_6",
        "Onshore_Wind_7",
        "Onshore_Wind_8",
        "Onshore_Wind_9",
    ]

    all_offwind = [
        "Offshore_Wind",
        "Offshore_Wind_1",
        "Offshore_Wind_10",
        "Offshore_Wind_11",
        "Offshore_Wind_12",
        "Offshore_Wind_13",
        "Offshore_Wind_14",
        "Offshore_Wind_15",
        "Offshore_Wind_2",
        "Offshore_Wind_3",
        "Offshore_Wind_4",
        "Offshore_Wind_5",
        "Offshore_Wind_6",
        "Offshore_Wind_7",
        "Offshore_Wind_8",
        "Offshore_Wind_9",
    ]

    all_onwind = [
        "Onshore_Wind",
        "Onshore_Wind_1",
        "Onshore_Wind_10",
        "Onshore_Wind_2",
        "Onshore_Wind_3",
        "Onshore_Wind_4",
        "Onshore_Wind_5",
        "Onshore_Wind_6",
        "Onshore_Wind_7",
        "Onshore_Wind_8",
        "Onshore_Wind_9",
    ]

    nrel_solar_PV = [
        "SolarDistUtil_PV_1",
        "SolarDistUtil_PV_2",
        "SolarDistUtil_PV_3",
        "SolarDistUtil_PV_4",
        "SolarDistUtil_PV_5",
        "SolarDistUtil_PV_6",
        "SolarDistUtil_PV_7",
        "SolarUtil_PV_1",
        "SolarUtil_PV_2",
        "SolarUtil_PV_3",
        "SolarUtil_PV_4",
        "SolarUtil_PV_5",
        "SolarUtil_PV_6",
        "SolarUtil_PV_7",
    ]

    nrel_offwind = [
        "Offshore_Wind_1",
        "Offshore_Wind_10",
        "Offshore_Wind_11",
        "Offshore_Wind_12",
        "Offshore_Wind_13",
        "Offshore_Wind_14",
        "Offshore_Wind_15",
        "Offshore_Wind_2",
        "Offshore_Wind_3",
        "Offshore_Wind_4",
        "Offshore_Wind_5",
        "Offshore_Wind_6",
        "Offshore_Wind_7",
        "Offshore_Wind_8",
        "Offshore_Wind_9",
    ]

    nrel_onwind = [
        "Onshore_Wind_1",
        "Onshore_Wind_10",
        "Onshore_Wind_2",
        "Onshore_Wind_3",
        "Onshore_Wind_4",
        "Onshore_Wind_5",
        "Onshore_Wind_6",
        "Onshore_Wind_7",
        "Onshore_Wind_8",
        "Onshore_Wind_9",
    ]

    #
    #
    # SETS
    agent = [
        "fossil_gen_agent",
        "renew_gen_agent",
        "transmission_agent",
        "demand_agent",
    ]
    prodn = ["fossil_gen_agent", "renew_gen_agent"]

    data = {
        "a": {"type": "set", "elements": agent, "text": "model agent type"},
        "prodn": {
            "type": "set",
            "domain": ["a"],
            "domain_info": "regular",
            "elements": prodn,
            "text": "model agent type that generated electricity",
        },
        "k": {"type": "set", "elements": k, "text": "all technolgy types in the model"},
        "gen": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": gen,
            "text": "all electricty generation technolgy types in the model",
        },
        "fossil": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": fossil,
            "text": "fossil technolgoies",
        },
        "hydro": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": hydro,
            "text": "hyrdo technologies",
        },
        "geothermal": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": geothermal,
            "text": "geothermal technolgoies",
        },
        "all_wind": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": all_wind,
            "text": "all wind technolgoies",
        },
        "all_offwind": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": all_offwind,
            "text": "all offshore wind technolgoies",
        },
        "all_onwind": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": all_onwind,
            "text": "all onshore wind technolgoies",
        },
        "all_solar": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": all_solar,
            "text": "all solar technolgoies",
        },
        "all_solar_PV": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": all_solar_PV,
            "text": "all solar PV technolgoies",
        },
        "all_distsolar_PV": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": all_distsolar_PV,
            "text": "all distributed solar PV technolgoies",
        },
        "all_utilsolar_PV": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": all_utilsolar_PV,
            "text": "all utility scale solar PV technolgoies",
        },
        "all_solar_therm": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": all_solar_therm,
            "text": "all utility scale solar thermal technolgoies",
        },
        "nuclear": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": nuclear,
            "text": "nuclear technologies",
        },
        "store": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": store,
            "text": "storage technologies",
        },
        "battery": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": battery,
            "text": "battery storage technologies",
        },
        "renew": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": renew,
            "text": "renewable technologies",
        },
        "nrel_offwind": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": nrel_offwind,
            "text": "offshore wind technologies from NREL",
        },
        "nrel_onwind": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": nrel_onwind,
            "text": "onshore wind technologies from NREL",
        },
        "nrel_solar_PV": {
            "type": "set",
            "domain": ["k"],
            "domain_info": "regular",
            "elements": nrel_solar_PV,
            "text": "solar PV technologies from NREL",
        },
    }

    # more set data
    reeds_region = list(set(reeds_reg["reeds.reg"].values))
    reeds_region.sort()

    reeds_balauth = list(set(reeds_reg["reeds.ba"].values))
    reeds_balauth.sort()

    uid = epa_needs.UniqueID_Final.tolist()

    regions = []
    regions.extend(latlng["fips"].unique().tolist())
    regions.extend(latlng["Region Name"].unique().tolist())
    regions.extend(latlng["state"].unique().tolist())
    regions.extend(reeds_region)
    regions.extend(reeds_balauth)

    fips = latlng["fips"].unique().tolist()
    ipm_regions = latlng["Region Name"].unique().tolist()

    states = latlng["state"].unique().tolist()
    states.sort()

    # time set data
    hrs = list(range(1, 8760 + 1))
    epoch = load.epoch.unique().tolist()
    time = list(zip(epoch, hrs))

    # MAPS -- mapping between geographic regions
    map_fips_reeds_regions = list(zip(reeds_reg["fips"], reeds_reg["reeds.reg"]))
    map_fips_reeds_balauth = list(zip(reeds_reg["fips"], reeds_reg["reeds.ba"]))
    map_fips_ipm = list(zip(latlng["fips"], latlng["Region Name"]))
    map_fips_state = list(zip(latlng["fips"], latlng["state"]))

    # MAPS -- locating generators to nodes
    map_uid_fips = list(zip(epa_needs["UniqueID_Final"], epa_needs["FIPS5"]))
    map_uid_ipm = list(zip(epa_needs["UniqueID_Final"], epa_needs["Region Name"]))
    map_uid_type = list(zip(epa_needs["UniqueID_Final"], epa_needs["PlantType"]))

    # PARAMETERS
    ldc_raw = dict(zip(zip(load.epoch, load.hour, load.Region), load.value))
    lat = dict(zip(latlng["fips"], latlng["lat"]))
    lng = dict(zip(latlng["fips"], latlng["lng"]))
    cap = dict(zip(epa_needs["UniqueID_Final"], epa_needs["Capacity (MW)"]))
    hr = dict(zip(epa_needs["UniqueID_Final"], epa_needs["Heat Rate (Btu/kWh)"]))
    pop = dict(zip(latlng.fips, latlng.population))

    # VRE related parameters
    nrel_solar = read_nrel_reeds_solar_data(to_csv=True)
    nrel_solar = dict(
        zip(
            zip(nrel_solar["tech"], nrel_solar["reeds_region"], nrel_solar["hour"]),
            nrel_solar["value"],
        )
    )

    nrel_wind = read_nrel_reeds_wind_data(to_csv=True)
    nrel_wind = dict(
        zip(
            zip(nrel_wind["tech"], nrel_wind["reeds_region"], nrel_wind["hour"]),
            nrel_wind["value"],
        )
    )

    nrel_wind_cap = pd.melt(
        nrel_wind_cap, id_vars=["reeds.reg", "trb", "tech"], var_name="cost_bin"
    )

    cost_bin = list(set(nrel_wind_cap["cost_bin"]))
    cost_bin.sort()

    nrel_wind_cap = dict(
        zip(
            zip(
                nrel_wind_cap["tech"],
                nrel_wind_cap["reeds.reg"],
                nrel_wind_cap["cost_bin"],
            ),
            nrel_wind_cap["value"],
        )
    )

    nrel_wind_cost = pd.melt(
        nrel_wind_cost, id_vars=["reeds.reg", "trb", "tech"], var_name="cost_bin"
    )
    nrel_wind_cost = dict(
        zip(
            zip(
                nrel_wind_cost["tech"],
                nrel_wind_cost["reeds.reg"],
                nrel_wind_cost["cost_bin"],
            ),
            nrel_wind_cost["value"],
        )
    )

    nrel_solar_cap = pd.melt(
        nrel_solar_cap, id_vars=["reeds.ba", "trb", "tech"], var_name="cost_bin"
    )
    nrel_solar_cap = dict(
        zip(
            zip(
                nrel_solar_cap["tech"],
                nrel_solar_cap["reeds.ba"],
                nrel_solar_cap["cost_bin"],
            ),
            nrel_solar_cap["value"],
        )
    )

    nrel_solar_cost = pd.melt(
        nrel_solar_cost, id_vars=["reeds.ba", "trb", "tech"], var_name="cost_bin"
    )
    nrel_solar_cost = dict(
        zip(
            zip(
                nrel_solar_cost["tech"],
                nrel_solar_cost["reeds.ba"],
                nrel_solar_cost["cost_bin"],
            ),
            nrel_solar_cost["value"],
        )
    )

    data_2 = {
        "reeds_regions": {
            "type": "set",
            "domain": ["regions"],
            "domain_info": "regular",
            "elements": reeds_region,
            "text": "NREL ReEDS regions",
        },
        "reeds_balauth": {
            "type": "set",
            "domain": ["regions"],
            "domain_info": "regular",
            "elements": reeds_balauth,
            "text": "NREL ReEDS balancing authority regions",
        },
        "uid": {
            "type": "set",
            "elements": uid,
            "text": "unique id numbers from EPA NEEDS database",
        },
        "regions": {"type": "set", "elements": regions, "text": "all regions"},
        "fips": {
            "type": "set",
            "domain": ["regions"],
            "domain_info": "regular",
            "elements": fips,
            "text": "all FIPS codes",
        },
        "ipm": {
            "type": "set",
            "domain": ["regions"],
            "domain_info": "regular",
            "elements": ipm_regions,
            "text": "all IPM codes",
        },
        "state": {
            "type": "set",
            "domain": ["regions"],
            "domain_info": "regular",
            "elements": states,
            "text": "all states",
        },
        "hrs": {"type": "set", "elements": hrs, "text": "hours in a year"},
        "epoch": {"type": "set", "elements": epoch, "text": "timestamp"},
        "time": {
            "type": "set",
            "domain": ["epoch", "hrs"],
            "domain_info": "regular",
            "elements": time,
            "text": "relationship between epoch and hrs",
        },
        "map_fips_reeds_regions": {
            "type": "set",
            "domain": ["regions", "regions"],
            "domain_info": "regular",
            "elements": map_fips_reeds_regions,
            "text": "map between FIPS5 codes and NREL ReEDS regions",
        },
        "map_fips_reeds_balauth": {
            "type": "set",
            "domain": ["regions", "regions"],
            "domain_info": "regular",
            "elements": map_fips_reeds_balauth,
            "text": "map between FIPS5 codes and NREL ReEDS balancing authority regions",
        },
        "map_fips_ipm": {
            "type": "set",
            "domain": ["regions", "regions"],
            "domain_info": "regular",
            "elements": map_fips_ipm,
            "text": "map between FIPS5 codes and IPM regions",
        },
        "map_fips_state": {
            "type": "set",
            "domain": ["regions", "regions"],
            "domain_info": "regular",
            "elements": map_fips_state,
            "text": "map between FIPS5 codes and State regions",
        },
        "map_uid_fips": {
            "type": "set",
            "domain": ["uid", "regions"],
            "domain_info": "regular",
            "elements": map_uid_fips,
            "text": "map between unit generator ID and FIPS5 codes",
        },
        "map_uid_ipm": {
            "type": "set",
            "domain": ["uid", "regions"],
            "domain_info": "regular",
            "elements": map_uid_ipm,
            "text": "map between unit generator ID and IPM region",
        },
        "map_uid_type": {
            "type": "set",
            "domain": ["uid", "k"],
            "domain_info": "regular",
            "elements": map_uid_type,
            "text": "map between unit generator ID and technolgy type",
        },
        "ldc_raw": {
            "type": "parameter",
            "domain": ["epoch", "hrs", "regions"],
            "domain_info": "regular",
            "elements": ldc_raw,
            "text": "raw load duration curves",
        },
        "lat": {
            "type": "parameter",
            "domain": ["regions"],
            "domain_info": "regular",
            "elements": lat,
            "text": "latitude",
        },
        "lng": {
            "type": "parameter",
            "domain": ["regions"],
            "domain_info": "regular",
            "elements": lng,
            "text": "longitude",
        },
        "cap": {
            "type": "parameter",
            "domain": ["uid"],
            "domain_info": "regular",
            "elements": cap,
            "text": "generation nameplate capacity (units: MW)",
        },
        "hr": {
            "type": "parameter",
            "domain": ["uid"],
            "domain_info": "regular",
            "elements": hr,
            "text": "generator efficiency (i.e., heat rate) (units: Btu/kWh)",
        },
        "population": {
            "type": "parameter",
            "domain": ["regions"],
            "domain_info": "regular",
            "elements": pop,
            "text": "population of a region (units: count)",
        },
        "nrel_solar_cf": {
            "type": "parameter",
            "domain": ["k", "regions", "hrs"],
            "domain_info": "regular",
            "elements": nrel_solar,
            "text": "solar capacity factor (units: unitless)",
        },
        "nrel_wind_cf": {
            "type": "parameter",
            "domain": ["k", "regions", "hrs"],
            "domain_info": "regular",
            "elements": nrel_wind,
            "text": "wind capacity factor (units: unitless)",
        },
        "cost_bin": {"type": "set", "elements": cost_bin, "text": "cost bin"},
        "nrel_wind_cap": {
            "type": "parameter",
            "domain": ["k", "regions", "cost_bin"],
            "domain_info": "regular",
            "elements": nrel_wind_cap,
            "text": "wind potential capacity (units: MW)",
        },
        "nrel_wind_cost": {
            "type": "parameter",
            "domain": ["k", "regions", "cost_bin"],
            "domain_info": "regular",
            "elements": nrel_wind_cost,
            "text": "wind costs (units: $/MW/year)",
        },
        "nrel_solar_cap": {
            "type": "parameter",
            "domain": ["k", "regions", "cost_bin"],
            "domain_info": "regular",
            "elements": nrel_solar_cap,
            "text": "solar potential capacity (units: MW)",
        },
        "nrel_solar_cost": {
            "type": "parameter",
            "domain": ["k", "regions", "cost_bin"],
            "domain_info": "regular",
            "elements": nrel_solar_cost,
            "text": "solar costs (units: $/MW/year)",
        },
    }

    #
    #
    # GDX creation
    gdx = gmsxfr.GdxContainer("/Applications/GAMS30.3/Resources/sysdir")

    # validata all data structures
    gdx.validate(data)
    gdx.validate(data_2)

    gdx.add_to_gdx(data, standardize_data=True, inplace=True, quality_checks=False)
    gdx.add_to_gdx(data_2, standardize_data=True, inplace=True, quality_checks=False)

    gdx.write_gdx("werewolf_data.gdx", compress=True)
